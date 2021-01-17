# This is my frist real world Program after finished my python cource  

import requests
import openpyxl
from bs4 import BeautifulSoup
import time

product_title=[]
discription=[]
imagelink=[]
rating=[]
def onepagecycle():
    #time.sleep(10)
    
    soup= BeautifulSoup(response.text,"html.parser")
    imagelist = soup.select(".image-fade_in_back")
    
    #print(imagelist)
    titlelist = soup.select(".title-wrapper")
    
    starrating=soup.select(".star-rating")
    
    for item in range(0,len(titlelist)):
        getstring = str(titlelist[item]) 
        length1= len('"<div class="title-wrapper"><p class="category uppercase is-smaller no-text-overflow product-cat op-7">"')
        product_name=getstring[length1-1:] 
        index2=product_name.index("<")
        product_title.append(product_name[0:index2]) # <-----------Product Name
        nextname = product_name[index2:]
        length3= len('</p><p class="name product-title woocommerce-loop-product__title"><a class="woocommerce-LoopProduct-link woocommerce-loop-product__link" href="https://plugintheme.net/shop/')
        laststr =(nextname[length3:])
        index3=laststr.index("<")
        disvript=(laststr[0:index3])
        discription.append(disvript.split('>', 1)[1]) # <------------------Discryption



        if ' 100w" width="247"/></noscript> </a></div>' in str(imagelist[item]):
            imglen=len(' 100w" width="247"/></noscript> </a></div>')
            totallen=len(imagelist[item])
            oneitem=str(imagelist[item])
            gotsummery=(oneitem[:totallen-(imglen+1)])
            linkbeforedit= str(gotsummery.rsplit(', ', 1)[1])
            remleng=len("-100x100.jpg")
            imagelink.append(linkbeforedit[0:len(linkbeforedit)-(remleng+1)] + ".jpg")
        elif ' 500w" width="247"/></noscript> </a></div>' in str(imagelist[item]):
            imglen=len(' 500w" width="247"/></noscript> </a></div>')
            totallen=len(imagelist[item])
            oneitem=str(imagelist[item])
            gotsummery=(oneitem[:totallen-(imglen+1)])
            linkbeforedit= str(gotsummery.rsplit(', ', 1)[1])
            #remleng=len("-100x100.jpg")
            #imagelink.append(linkbeforedit[0:len(linkbeforedit)-(remleng+1)] + ".jpg")
            imagelink.append(linkbeforedit)

    for items in range(0,len(starrating)):
        beforenumb=starrating[items].select(".rating")
        len12=len('<strong class="rating">')
        itemtotext=str(beforenumb)
        t1= itemtotext[len12+1:]
        ind=t1.index("<")
        rating.append(t1[0:ind])
#-------------------------------------------------------
# print(len(rating))
# print(len(discription))
# print(len(product_title))
# print(imagelink)
# Test code for all pages that how works 
response = requests.get("https://plugintheme.net/shop/page/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/2/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/3/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/4/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/5/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/6/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/7/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/8/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/9/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/10/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/11/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/12/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/13/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/14/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/15/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/16/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/17/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/18/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/19/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/20/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/21/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/22/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/23/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/24/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/25/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/26/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/27/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/28/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/29/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/30/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/31/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/32/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/33/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/34/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/34/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/35/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/36/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/37/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/38/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/39/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/40/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/41/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/42/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/43/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/44/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/45/")
onepagecycle()
response = requests.get("https://plugintheme.net/shop/page/46/")
onepagecycle()


wb = openpyxl.load_workbook("Extracted_Data.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
#cell = wb.cell(raw=1,column=1)

for cells in range(1,len(discription)):
    sheet.append([str(discription[cells]),str(imagelink[cells])])
wb.save("Extracted_Data.xlsx")
# all 46 pages at the web site
# #response = requests.get("https://plugintheme.net/shop/")

